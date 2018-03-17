#! python3
# This program scrapes thebeerstore.ca for a list of url's to every beer
# they sell. The beers are categorized into Ale, Lager, Malt & Stout.

import requests, bs4, re, openpyxl, lxml, beer_store_scraper, threading, queue
import insertion_sort as sorting
from openpyxl import Workbook

q = queue.Queue()
beerTypes = ['Lager', 'Ale', 'Malt', 'Stout']
threads = []
beerList = []

for i, beerType in enumerate(beerTypes):
    newThread = threading.Thread(target=beer_store_scraper.scrape, args=(beerType, q))
    threads.append(newThread)
    newThread.start()

for i in range(len(threads)):
    beerList.append(q.get())
    q.task_done()

q.join()
for currentThread in threads:
    currentThread.join()
# Each thread now finished. All scraping has been completed

# Change beerList from a 3-D array to a 2-D array
for i in range(3):
    for j in range(len(beerList[1])):
        beerList[0].append(beerList[1][j])
    del beerList[1]
beerList = beerList[0]
# Beer list now formatted properly. Verified by testing.

# Get price per drink for each beer (17.7441 mL alcohol/drink)
mL_per_drink = 17.7441 # mL
for i in range(len(beerList)):
    alcoholVol = float(beerList[i][2]) * float(beerList[i][3]) * float(beerList[i][4])
    dollars_per_mL = float(beerList[i][5]) / alcoholVol
    costPerDrink = dollars_per_mL * mL_per_drink
    beerList[i].append(costPerDrink)

# Sort beerList by price per drink
beerList = sorting.sort(beerList)

# Create a worksheet
wb = Workbook()
ws = wb.active
ws.title = 'Sorted By Price'
ws['B3'] = 'Beer'
ws['C3'] = 'Sold as'
ws['D3'] = 'Quantity'
ws['E3'] = 'Volume (ml)'
ws['F3'] = 'ABV'
ws['G3'] = 'Price'
ws['G3'].number_format = '$#.##' # not working need to find out why
ws['H3'] = 'Price per Drink'
ws['H3'].number_format = '$#.##'
ws.column_dimensions['B'].width = 27
ws.column_dimensions['C'].width = 20
ws.column_dimensions['H'].width = 12

row = 4
# Print to spreadsheet
for i in range(len(beerList)):
    nameIndex = 'B' + str(row)
    sizeIndex = 'C' + str(row)
    quantityIndex = 'D' + str(row)
    volIndex = 'E' + str(row)
    abvIndex = 'F' + str(row)
    priceIndex = 'G' + str(row)
    costPerDrinkIndex = 'H' + str(row)
    row += 1

    # Import list into Excel
    ws[nameIndex] = beerList[i][0]
    ws[sizeIndex] = beerList[i][1]
    ws[quantityIndex] = int(beerList[i][2])
    ws[volIndex] = float(beerList[i][3])
    ws[abvIndex] = beerList[i][4]
    ws[priceIndex] = float(beerList[i][5])
    ws[costPerDrinkIndex] = round(beerList[i][6],2)

wb.save('beer_store_catalog_threaded.xlsx')
