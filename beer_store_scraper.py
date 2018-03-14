#! python3
# This program scrapes thebeerstore.ca for a list of url's to every beer
# they sell. The beers are categorized into Ale, Lager, Malt & Stout.

import requests, bs4, re, openpyxl, lxml
from openpyxl import Workbook
from openpyxl.styles import Font

def scrape(beerType):

    # Get all links for the type of beer
    url = 'http://www.thebeerstore.ca/beers/search/beer_type--' + beerType
    linksRequest = requests.get(url)
    linksText = bs4.BeautifulSoup(linksRequest.text, "lxml")
    rawLinks = linksText.select('.brand-link')

    # Regular expression declarations
    sizeRegex = re.compile(r'\d+.*ml')
    volumeRegex = re.compile(r'\d+\sml')
    quantityRegex = re.compile(r'\d+')
    priceRegex = re.compile(r'\d+\.\d{2}')
    linkRegex = re.compile(r'/beers/\w+(-\%\w+(-\w+)*)?(-\w+)*')
    nameRegex = re.compile(r'>.*<')
    ABVRegex = re.compile(r'\d+\.\d')
    beerList = []

    for i in range(len(rawLinks)):
        # Get the webpage of this specifc link
        rawLinksString = str(rawLinks[i])
        match = linkRegex.search(rawLinksString)
        fullLink = 'http://www.thebeerstore.ca' + match.group(0)
        beerInfo = requests.get(fullLink)
        beerSoup = bs4.BeautifulSoup(beerInfo.text, "lxml")

        # Get prices and sizes of the beer
        prices = beerSoup.select('.price')
        sizes = beerSoup.select('.size')
        
        # Get the name of the beer
        names = beerSoup.select('.page-title')
        nameMatch = nameRegex.search(str(names[0]))
        name = nameMatch.group(0)[1:-1]

        # Get the ABV of the beer
        abvContent = beerSoup.select('dd')
        abvMatch = ABVRegex.search(str(abvContent))
        ABV = float(abvMatch.group(0)) / 100

        # Iterate through each size option to get the price, size, vol, and quantity of each
        for j in range(len(prices)):
            price = priceRegex.search(str(prices[j])).group(0)
            volume = volumeRegex.search(str(sizes[j])).group(0)
            quantity = quantityRegex.search(str(sizes[j])).group(0)
            size = sizeRegex.search(str(sizes[j])).group(0)
            beerList.append([name, size, quantity, volume[:-3], ABV, price]) # remove ' ml'

    return beerList

beerTypes = ['Lagers', 'Ales', 'Malts', 'Stouts']
lagerList = scrape('Lager')
aleList = scrape('Ale')
maltList = scrape('Malt')
stoutList = scrape('Stout')
beerLists = [lagerList, aleList, maltList, stoutList]

# Create 4 worksheets
wb = Workbook()
ws = wb.active
ws.title = 'Lagers'
wb.create_sheet(title='Ales')
wb.create_sheet(title='Malts')
wb.create_sheet(title='Stouts')

bold11Font = Font(size=11, bold=True)
mL_per_drink = 17.7441 # ml

for counter, beerList in enumerate(beerLists):
    row = 4
    
    # Set headers and change active sheet
    ws = wb[beerTypes[counter]]
    ws['B3'] = 'Beer'
    #ws['B3'].font = bold11Font
    ws['C3'] = 'Sold as'
    #ws['C3'].font = bold11Font
    ws['D3'] = 'Quantity'
    #ws['D3'].font = bold11Font
    ws['E3'] = 'Volume (ml)'
    #ws['E3'].font = bold11Font
    ws['F3'] = 'ABV'
    #ws['F3'].font = bold11Font
    ws['G3'] = 'Price'
    #ws['G3'].font = bold11Font
    ws['G3'].number_format = '$#.##'
    ws['H3'] = 'Price per Drink'
    #ws['H3'].font = bold11Font
    ws['H3'].number_format = '$#.##'
    ws.column_dimensions['B'].width = 27
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['H'].width = 12
    
    for i in range(len(beerList)):
        nameIndex = 'B' + str(row)
        sizeIndex = 'C' + str(row)
        quantityIndex = 'D' + str(row)
        volIndex = 'E' + str(row)
        abvIndex = 'F' + str(row)
        priceIndex = 'G' + str(row)
        costPerDrinkIndex = 'H' + str(row)
        row += 1

        # Import list into Excel
        ws[nameIndex] = beerList[i][0]
        ws[sizeIndex] = beerList[i][1]
        ws[quantityIndex] = int(beerList[i][2])
        ws[volIndex] = float(beerList[i][3])
        ws[abvIndex] = beerList[i][4]
        ws[priceIndex] = float(beerList[i][5])
        
        # Get price per drink (17.7441 mL alcohol)
        alcoholVol = float(beerList[i][2]) * float(beerList[i][3]) * float(beerList[i][4])
        dollars_per_mL = float(beerList[i][5]) / alcoholVol
        costPerDrink = dollars_per_mL * mL_per_drink
        ws[costPerDrinkIndex] = round(costPerDrink,2)

wb.save('beer_store_catalog.xlsx')
